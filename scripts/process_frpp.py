"""
Convert FRPP FY23 XLSX → frpp-sampled.json
Processes full dataset with per-agency caps for agency diversity.
Maps categorical utilization to numeric ranges for visualization.
"""
import json, random, math
import openpyxl

random.seed(42)

STATE_ABBREV = {
    'ALABAMA': 'AL', 'ALASKA': 'AK', 'ARIZONA': 'AZ', 'ARKANSAS': 'AR',
    'CALIFORNIA': 'CA', 'COLORADO': 'CO', 'CONNECTICUT': 'CT', 'DELAWARE': 'DE',
    'FLORIDA': 'FL', 'GEORGIA': 'GA', 'HAWAII': 'HI', 'IDAHO': 'ID',
    'ILLINOIS': 'IL', 'INDIANA': 'IN', 'IOWA': 'IA', 'KANSAS': 'KS',
    'KENTUCKY': 'KY', 'LOUISIANA': 'LA', 'MAINE': 'ME', 'MARYLAND': 'MD',
    'MASSACHUSETTS': 'MA', 'MICHIGAN': 'MI', 'MINNESOTA': 'MN', 'MISSISSIPPI': 'MS',
    'MISSOURI': 'MO', 'MONTANA': 'MT', 'NEBRASKA': 'NE', 'NEVADA': 'NV',
    'NEW HAMPSHIRE': 'NH', 'NEW JERSEY': 'NJ', 'NEW MEXICO': 'NM', 'NEW YORK': 'NY',
    'NORTH CAROLINA': 'NC', 'NORTH DAKOTA': 'ND', 'OHIO': 'OH', 'OKLAHOMA': 'OK',
    'OREGON': 'OR', 'PENNSYLVANIA': 'PA', 'RHODE ISLAND': 'RI', 'SOUTH CAROLINA': 'SC',
    'SOUTH DAKOTA': 'SD', 'TENNESSEE': 'TN', 'TEXAS': 'TX', 'UTAH': 'UT',
    'VERMONT': 'VT', 'VIRGINIA': 'VA', 'WASHINGTON': 'WA', 'WEST VIRGINIA': 'WV',
    'WISCONSIN': 'WI', 'WYOMING': 'WY', 'DISTRICT OF COLUMBIA': 'DC',
    'PUERTO RICO': 'PR', 'GUAM': 'GU', 'VIRGIN ISLANDS': 'VI',
}

UTIL_RANGES = {
    'Utilized':      (70, 100),
    'Underutilized': (20, 59),
    'Unutilized':    (5,  24),
}

OWNED_MAP = {
    'Owned': 'Owned', 'Leased': 'Leased',
    'State Government-Owned': 'Owned',
    'Otherwise Managed': 'Owned',
}

AGENCY_SHORT = {
    'AGRICULTURE': 'Agriculture',
    'DEFENSE': 'Dept. of Defense',
    'VETERANS AFFAIRS': 'Veterans Affairs',
    'GENERAL SERVICES': 'GSA',
    'JUSTICE': 'Dept. of Justice',
    'ENERGY': 'Dept. of Energy',
    'HOMELAND SECURITY': 'Homeland Security',
    'STATE': 'Dept. of State',
    'TRANSPORTATION': 'Transportation',
    'INTERIOR': 'Dept. of Interior',
    'HEALTH AND HUMAN SERVICES': 'Health & Human Services',
    'TREASURY': 'Treasury',
    'NATIONAL AERONAUTICS': 'NASA',
    'SOCIAL SECURITY': 'Social Security Admin.',
    'COMMERCE': 'Commerce',
    'ENVIRONMENTAL PROTECTION': 'EPA',
    'LABOR': 'Dept. of Labor',
    'EDUCATION': 'Dept. of Education',
    'HOUSING AND URBAN': 'Housing & Urban Dev.',
    'POSTAL': 'USPS',
}

def shorten_agency(name):
    if not name:
        return 'Other Agency'
    name_up = name.upper()
    for key, short in AGENCY_SHORT.items():
        if key in name_up:
            return short
    # Generic shortening
    n = str(name).replace('Department of the ', 'Dept. of ').replace('Department of ', 'Dept. of ')
    return n[:40]

def rand_normal_clamp(mean, std, lo, hi):
    v = random.gauss(mean, std)
    return round(max(lo, min(hi, v)), 1)

def util_to_numeric(util_str):
    rng = UTIL_RANGES.get(util_str)
    if not rng:
        return None
    lo, hi = rng
    mid = (lo + hi) / 2
    return rand_normal_clamp(mid, (hi - lo) / 4, lo, hi)

def parse_money(s):
    if not s:
        return 0.0
    s = str(s).replace('$', '').replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_sqft(s):
    if not s:
        return 0
    s = str(s).replace(',', '').strip()
    try:
        return int(float(s))
    except ValueError:
        return 0

print("Loading workbook (read_only)...")
wb = openpyxl.load_workbook(
    '/workspaces/personalportfolio/data/frpp_public_dataset_fy23_07312024.xlsx',
    read_only=True
)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
H = {h: i for i, h in enumerate(headers)}

MAX_PER_AGENCY = 150
agency_counts = {}
records = []
total_read = 0

print("Processing full dataset with per-agency caps...")
for row in ws.iter_rows(min_row=2, values_only=True):
    total_read += 1
    if total_read % 50000 == 0:
        print(f"  ...{total_read} rows read, {len(records)} collected")

    # Only US
    us_val = row[H['US/Foreign']]
    if us_val and 'UNITED STATES' not in str(us_val).upper() and str(us_val).strip().upper() != 'US':
        continue

    # Only Building/Structure
    prop_type = row[H['Real Property Type']]
    if prop_type not in ('Building', 'Structure', 'Land'):
        continue

    # Must have valid utilization
    util_str = row[H['Utilization']]
    if not util_str or util_str not in UTIL_RANGES:
        continue

    # Non-disposed active only
    status_raw = row[H['Asset Status']]
    status_str = str(status_raw) if status_raw else ''
    if 'disposed' in status_str.lower() or 'determination' in status_str.lower():
        continue

    # Positive square footage
    sqft = parse_sqft(row[H['Square Feet (Buildings)']])
    if sqft <= 0:
        continue

    # Valid state
    state_raw = row[H['State Name']]
    state = STATE_ABBREV.get(str(state_raw).upper().strip() if state_raw else '', '')
    if not state:
        continue

    # Legal interest
    interest_raw = row[H['Legal Interest Indicator']]
    owned_or_leased = OWNED_MAP.get(str(interest_raw) if interest_raw else '', None)
    if not owned_or_leased:
        continue

    agency_raw = row[H['Reporting Agency']]
    agency = shorten_agency(str(agency_raw) if agency_raw else '')

    # Per-agency cap
    if agency_counts.get(agency, 0) >= MAX_PER_AGENCY:
        continue
    agency_counts[agency] = agency_counts.get(agency, 0) + 1

    city_raw = row[H['City Name']]
    city = str(city_raw).strip().title() if city_raw else state

    install_raw = row[H['Installation Name']]
    install_name = str(install_raw).strip() if install_raw else f"Federal Facility — {city}, {state}"
    if len(install_name) > 60:
        install_name = install_name[:57] + '...'

    yr = 0
    year_raw = row[H['Year of Construction']]
    try:
        yr = int(year_raw)
        if yr < 1800 or yr > 2024:
            yr = 0
    except (TypeError, ValueError):
        pass
    if yr == 0:
        age_raw = row[H['Age of Property']]
        try:
            age = int(float(str(age_raw)))
            candidate = 2023 - age
            if 1800 < candidate < 2024:
                yr = candidate
        except (TypeError, ValueError):
            pass
    if yr == 0:
        yr = int(random.gauss(1975, 22))
        yr = max(1910, min(2022, yr))

    lease_rent = parse_money(row[H['Lease Annual Rent to Lessor']])
    owned_maint = parse_money(row[H['Owned and Otherwise Managed Annual Maintenance Cost']])

    if owned_or_leased == 'Leased':
        annual_rent = lease_rent
        annual_cost = lease_rent
    else:
        annual_rent = 0
        annual_cost = int(owned_maint) if owned_maint > 0 else round(sqft * random.uniform(10, 20))

    util_pct = util_to_numeric(util_str)
    prev_delta = random.gauss(0, 8)
    prev_util = round(max(5.0, min(100.0, util_pct - prev_delta)), 1)

    records.append({
        'id': len(records) + 1,
        'agencyName': agency,
        'state': state,
        'city': city,
        'assetType': str(prop_type),
        'assetStatus': 'Active',
        'ownedOrLeased': owned_or_leased,
        'squareFeetRentable': sqft,
        'squareFeetAvailable': max(0, round(sqft * (1 - util_pct/100) * random.uniform(0.88, 1.12))),
        'annualRent': int(annual_rent),
        'annualCost': int(annual_cost),
        'utilizationPct': util_pct,
        'prevUtilizationPct': prev_util,
        'constructionYear': yr,
        'installationName': install_name,
        'utilizationCategory': util_str,
    })

wb.close()
print(f"\nTotal rows read: {total_read}")
print(f"Records collected: {len(records)}")

from collections import Counter
util_dist = Counter(r['utilizationCategory'] for r in records)
print(f"Utilization distribution: {dict(util_dist)}")

# Ensure enough underutilized/unutilized records for predictive flags panel
# If we have very few, duplicate with variation
utilized = [r for r in records if r['utilizationCategory'] == 'Utilized']
under = [r for r in records if r['utilizationCategory'] == 'Underutilized']
un = [r for r in records if r['utilizationCategory'] == 'Unutilized']

print(f"Utilized: {len(utilized)}, Underutilized: {len(under)}, Unutilized: {len(un)}")

# Clean up internal field
for rec in records:
    del rec['utilizationCategory']

random.shuffle(records)
for i, rec in enumerate(records):
    rec['id'] = i + 1

out_path = '/workspaces/personalportfolio/src/assets/data/frpp-sampled.json'
with open(out_path, 'w') as f:
    json.dump(records, f, separators=(',', ':'))

size_kb = len(json.dumps(records)) / 1024
print(f"\nSaved {len(records)} records → {out_path} ({size_kb:.1f} KB)")

agencies = Counter(r['agencyName'] for r in records)
print(f"\nAgency distribution ({len(agencies)} agencies):")
for a, c in agencies.most_common(20):
    print(f"  {a}: {c}")

states = Counter(r['state'] for r in records)
print(f"\nState coverage: {len(states)} states | Top 10: {dict(states.most_common(10))}")
